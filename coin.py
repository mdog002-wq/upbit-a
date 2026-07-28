import os
import time
import datetime
import smtplib
import requests
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import pandas as pd
import pyupbit
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# [설정] GitHub Secrets에서 이메일 정보 가져오기 (보안 적용)
# ==============================================================================
SENDER_EMAIL = os.environ.get("mdog002@gmail.com")       # mdog002@gmail.com
EMAIL_PASSWORD = os.environ.get("alub bzok obsb vfmx")   # alub bzok obsb vfmx
RECEIVER_EMAIL = os.environ.get("mdog002@gmail.com")   # mdog002@gmail.com
EXCEL_FILE_PATH = "업비트_원화마켓_매집점수_날짜별기록.xlsx"

# ==============================================================================
# [유틸] 업비트 '원화(KRW) 마켓' 전 종목 및 한글명 가져오기
# ==============================================================================
def get_krw_upbit_tickers():
    url = "https://api.upbit.com/v1/market/all"
    headers = {"accept": "application/json"}
    
    try:
        response = requests.get(url, headers=headers)
        data = response.json()
        
        krw_coins = []
        for coin in data:
            # 원화(KRW) 마켓 코인만 엄격히 추출
            if coin['market'].startswith("KRW-"):
                krw_coins.append({
                    'ticker': coin['market'],
                    'korean_name': coin['korean_name'],
                    'english_name': coin['english_name'],
                    'symbol': coin['market'].replace("KRW-", "")
                })
        return krw_coins
    except Exception as e:
        print(f"❌ 업비트 원화 코인 목록 조회 실패: {e}")
        return []

# ==============================================================================
# [알고리즘] 매집 점수(100점 만점) 산출
# ==============================================================================
def calculate_score(vol_ratio, trade_value_krw, price_change, is_yangbong, ma_gap):
    score = 0
    
    # 1. 거래량 급증 점수 (최대 35점)
    if vol_ratio >= 4.0: score += 35
    elif vol_ratio >= 2.5: score += 28
    elif vol_ratio >= 1.8: score += 20
    elif vol_ratio >= 1.3: score += 12
    else: score += 5

    # 2. 거래대금 규모 점수 (최대 25점)
    trade_value_eow = trade_value_krw / 100_000_000
    if trade_value_eow >= 50: score += 25
    elif trade_value_eow >= 20: score += 20
    elif trade_value_eow >= 10: score += 15
    elif trade_value_eow >= 5: score += 10
    else: score += 5

    # 3. 가격 상승률 및 양봉 여부 (최대 20점)
    if is_yangbong and 0.3 <= price_change <= 7.0: score += 20
    elif is_yangbong and price_change > 7.0: score += 12
    elif not is_yangbong: score += 0

    # 4. 이동평균선 수렴도 (최대 20점)
    if ma_gap <= 3.0: score += 20
    elif ma_gap <= 6.0: score += 15
    elif ma_gap <= 10.0: score += 10
    else: score += 3

    return score

# ==============================================================================
# [분석 엔진] 원화 마켓 전 종목 완전 수집 (누락 방지 재시도 로직)
# ==============================================================================
def scan_and_rank_coins():
    print("--------------------------------------------------")
    print("🚀 업비트 원화(KRW) 마켓 전 종목 전수 조사 시작...")
    
    krw_coins = get_krw_upbit_tickers()
    if not krw_coins:
        return pd.DataFrame()

    print(f"총 {len(krw_coins)}개 원화 마켓 종목 분석 중...\n")
    results = []
    
    for item in krw_coins:
        ticker = item['ticker']
        korean_name = item['korean_name']
        symbol = item['symbol']
        
        # API 조회 재시도(Retry) 로직 (최대 3회)
        df_daily = None
        for retry in range(3):
            try:
                time.sleep(0.06)
                df_daily = pyupbit.get_ohlcv(ticker, interval="day", count=60)
                if df_daily is not None and not df_daily.empty:
                    break
            except Exception:
                time.sleep(0.1)

        # 데이터 조회가 안 되는 특수 종목(신규 상장/정지 등) 예외 수집 처리
        if df_daily is None or df_daily.empty:
            results.append({
                "코인명": korean_name,
                "심볼": symbol,
                "매집점수": 0,
                "당일 변동률(%)": 0.0,
                "거래량 급증(배)": 0.0,
                "거래대금(억원)": 0.0,
                "현재가(KRW)": 0,
                "이평선 수렴도(%)": 0.0,
                "비고": "OHLCV 데이터 미제공(신규/정지)"
            })
            print(f"  [예외 수집] {korean_name}({symbol}) - 데이터 미제공 포함")
            continue

        try:
            latest = df_daily.iloc[-1]
            
            # 신규 상장 등으로 일봉 개수가 부족할 때 안전 처리
            if len(df_daily) >= 21:
                prev_20_vol_avg = df_daily['volume'].iloc[-21:-1].mean()
            else:
                prev_20_vol_avg = df_daily['volume'].mean()

            vol_ratio = (latest['volume'] / prev_20_vol_avg) if prev_20_vol_avg > 0 else 0
            trade_value_krw = latest['value']
            
            if len(df_daily) >= 2:
                price_change = ((latest['close'] - df_daily['close'].iloc[-2]) / df_daily['close'].iloc[-2]) * 100
            else:
                price_change = 0.0
                
            is_yangbong = latest['close'] >= latest['open']
            
            ma20 = df_daily['close'].rolling(20).mean().iloc[-1] if len(df_daily) >= 20 else latest['close']
            ma60 = df_daily['close'].rolling(60).mean().iloc[-1] if len(df_daily) >= 60 else ma20
            ma_gap = abs(ma20 - ma60) / ma20 * 100 if ma20 > 0 else 0

            total_score = calculate_score(vol_ratio, trade_value_krw, price_change, is_yangbong, ma_gap)
            
            results.append({
                "코인명": korean_name,
                "심볼": symbol,
                "매집점수": total_score,
                "당일 변동률(%)": round(price_change, 2),
                "거래량 급증(배)": round(vol_ratio, 2),
                "거래대금(억원)": round(trade_value_krw / 100_000_000, 1),
                "현재가(KRW)": latest['close'],
                "이평선 수렴도(%)": round(ma_gap, 2),
                "비고": "정상"
            })
            print(f"  [수집 완료] {korean_name}({symbol}) | 매집점수: {total_score}점")
                
        except Exception as e:
            results.append({
                "코인명": korean_name,
                "심볼": symbol,
                "매집점수": 0,
                "당일 변동률(%)": 0.0,
                "거래량 급증(배)": 0.0,
                "거래대금(억원)": 0.0,
                "현재가(KRW)": 0,
                "이평선 수렴도(%)": 0.0,
                "비고": f"연산 오류({e})"
            })
            continue
            
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values(by="매집점수", ascending=False)
    return df

# ==============================================================================
# [날짜별 엑셀 시트 저장]
# ==============================================================================
def save_daily_excel_sheet(df):
    if df.empty:
        return None

    now = datetime.datetime.now()
    sheet_name = now.strftime("%Y-%m-%d_%H시")
    
    if os.path.exists(EXCEL_FILE_PATH):
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        wb = openpyxl.Workbook()
        
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        
    ws = wb.create_sheet(title=sheet_name)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="맑은 고딕", size=10)
    high_score_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row in range(2, ws.max_row + 1):
        score_cell = ws.cell(row=row, column=3) # 매집점수 열 위치 (3번째 열)
        is_high_score = score_cell.value and float(score_cell.value) >= 70
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col != 7 else "right", vertical="center")
            
            if is_high_score:
                cell.fill = high_score_fill
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    wb.save(EXCEL_FILE_PATH)
    wb.close()
    
    print("--------------------------------------------------")
    print(f"📁 [{sheet_name}] 시트에 총 {len(df)}개 원화 코인 수집 완료!")
    return EXCEL_FILE_PATH

# ==============================================================================
# [이메일 첨부 발송]
# ==============================================================================
def send_email_with_excel(file_path):
    if not file_path or not os.path.exists(file_path):
        return

    if not SENDER_EMAIL or not EMAIL_PASSWORD or not RECEIVER_EMAIL:
        print("❌ 환경변수(Secrets) 설정에 문제가 있어 이메일을 발송하지 못했습니다.")
        return

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    msg = MIMEMultipart()
    msg["Subject"] = f"📊 [원화마켓 전종목] 업비트 매집 점수 리포트 ({now_str})"
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECEIVER_EMAIL
    
    body = f"""안녕하세요.

업비트 원화(KRW) 마켓 전체 상장 종목 분석이 완료되었습니다.

• 분석 시각: {now_str}
• 분석 대상: 원화 마켓 전체 종목

자세한 수치는 첨부된 엑셀 파일을 확인해 주세요.
"""
    msg.attach(MIMEText(body, "plain", "utf-8"))

    try:
        with open(file_path, "rb") as f:
            part = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
        file_name = os.path.basename(file_path)
        part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", file_name))
        msg.attach(part)
    except Exception as e:
        print(f"❌ 파일 첨부 실패: {e}")
        return

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, EMAIL_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
        server.quit()
        print(f"📧 엑셀 첨부 이메일 발송 성공! (수신: {RECEIVER_EMAIL})")
        print("--------------------------------------------------")
    except Exception as e:
        print(f"❌ 이메일 발송 실패: {e}")

if __name__ == "__main__":
    df_ranked = scan_and_rank_coins()
    excel_file = save_daily_excel_sheet(df_ranked)
    send_email_with_excel(excel_file)